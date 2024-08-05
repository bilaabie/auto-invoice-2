# excel_utils.py
from openpyxl import load_workbook
from datetime import datetime
from config import EXCEL_FILE

def input_to_excel2(invoice_number, date, due_date, total, bill_to):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws = wb['Sheet1']
    table = ws.tables['Table1']
    next_row = ws.max_row + 1

    if ws.cell(next_row - 1, 1).value is None:
        row_number = 1
    else:
        row_number = ws.cell(next_row - 1, 1).value + 1

    ws.cell(next_row, 1, row_number)
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.cell(next_row, 8, current_timestamp)

    ws.cell(next_row, 2, invoice_number)
    ws.cell(next_row, 3, date)
    ws.cell(next_row, 4, due_date)
    ws.cell(next_row, 5, total)
    ws.cell(next_row, 6, bill_to)
    ws.cell(next_row, 7, 'Pending')

    table.ref = f"{table.ref.split(':')[0]}:{ws.cell(next_row, ws.max_column).coordinate}"
    wb.save(EXCEL_FILE)

def update_excel_status(invoice_number, status):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws = wb['Sheet1']

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if row[1].value == invoice_number:
            row[6].value = status
            break

    wb.save(EXCEL_FILE)

def send_approval_request(invoice_number, date, due_date, total, bill_to):
    from config import WEBHOOK_URL
    adaptive_card_payload = {
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "themeColor": "0076D7",
        "summary": "Approval Request",
        "sections": [
            {
                "activityTitle": "Approval Request",
                "activitySubtitle": "Please approve or reject the following request.",
                "facts": [
                    {"name": "Invoice Number", "value": invoice_number},
                    {"name": "Date", "value": date},
                    {"name": "Due Date", "value": due_date},
                    {"name": "Total", "value": total},
                    {"name": "Bill to", "value": bill_to},
                ],
                "potentialAction": [
                    {
                        "@type": "HttpPOST",
                        "name": "Approve",
                        "target": f"{public_url}/webhook",
                        "body": json.dumps({"action": "Approve", "invoice_number": invoice_number})
                    },
                    {
                        "@type": "HttpPOST",
                        "name": "Reject",
                        "target": f"{public_url}/webhook",
                        "body": json.dumps({"action": "Reject", "invoice_number": invoice_number})
                    }
                ]
            }
        ]
    }
    response = requests.post(WEBHOOK_URL, json=adaptive_card_payload)
    print(response.status_code, response.text)
